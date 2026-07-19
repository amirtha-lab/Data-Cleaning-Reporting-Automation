import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# 1. SIMULATE MESSY RAW DATA
# ==========================================
# In reality, you would load this using: df = pd.read_csv("your_data.csv")
raw_data = {
    "Transaction_ID": [101, 102, 103, 103, 104, 105, 106],
    "Date": ["2026-07-15", "2026/07/16", "2026-07-17", "2026-07-17", "18-07-2026", None, "2026-07-19"],
    "Region": ["North", "north", "South", "South", "East", "West", "north"],
    "Revenue": [1500, np.nan, 2100, 2100, -500, 3200, 1800],
    "Units_Sold": [10, 5, np.nan, np.nan, 3, 15, 12]
}
df = pd.DataFrame(raw_data)
print("--- RAW DATA ---")
print(df, "\n")

# ==========================================
# 2. DATA CLEANING WORKFLOW
# ==========================================

# A. Handle Duplicates
df = df.drop_duplicates(subset=["Transaction_ID"], keep="first")

# B. Standardize Text & Inconsistencies (Fix casing)
df["Region"] = df["Region"].str.strip().str.capitalize()

# C. Standardize Dates
df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=False)
# Fill missing dates with a placeholder or forward fill
df["Date"] = df["Date"].fillna(method="ffill").fillna(pd.Timestamp.now().normalize())

# D. Handle Missing & Invalid Numeric Values
# Fix negative revenue (assuming input error)
df["Revenue"] = df["Revenue"].abs()
# Impute missing Revenue with the median of that region
df["Revenue"] = df.groupby("Region")["Revenue"].transform(lambda x: x.fillna(x.median()))

# Impute missing Units Sold based on average price per unit if possible, or fallback to 1
df["Units_Sold"] = df["Units_Sold"].fillna(1).astype(int)

print("--- CLEANED DATA ---")
print(df, "\n")

# ==========================================
# 3. AGGREGATION & REPORTING
# ==========================================
# Create a summary table for the report
summary_df = df.groupby("Region").agg(
    Total_Revenue=("Revenue", "sum"),
    Total_Units=("Units_Sold", "sum"),
    Avg_Transaction_Value=("Revenue", "mean")
).reset_index()

# ==========================================
# 4. EXCEL AUTOMATION & STYLING
# ==========================================
wb = Workbook()

# --- Sheet 1: Dashboard / Summary ---
ws1 = wb.active
ws1.title = "Summary Report"
ws1.views.sheetView[0].showGridLines = True

# Title Block
ws1["A1"] = "Executive Performance Report"
ws1["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
ws1["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws1.merge_cells("A1:D1")
ws1.row_dimensions[1].height = 40
ws1["A1"].alignment = Alignment(vertical="center", indent=1)

# Write Summary Data
ws1.append([]) # Blank row
ws1.append(["Region", "Total Revenue", "Total Units Sold", "Avg Order Value"])

for r in dataframe_to_rows(summary_df, index=False, header=False):
    ws1.append(r)

# Format Summary Table
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="000000")
thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

for row in ws1.iter_rows(min_row=3, max_row=3+len(summary_df), min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        if cell.row == 3:
            cell.fill = header_fill
            cell.font = header_font
        else:
            # Format numbers
            if cell.column in [2, 4]: # Revenue columns
                cell.number_format = '$#,##0.00'
            elif cell.column == 3: # Units
                cell.number_format = '#,##0'

# --- Sheet 2: Cleaned Data Data ---
ws2 = wb.create_sheet(title="Cleaned Data")
ws2.views.sheetView[0].showGridLines = True

# Write Dataframe to Sheet 2
for r in dataframe_to_rows(df, index=False, header=True):
    ws2.append(r)

# Auto-fit columns across both sheets
for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save the workbook
output_filename = "Automated_Sales_Report.xlsx"
wb.save(output_filename)
print(f"🎉 Success! Report saved as '{output_filename}' in your current directory.")